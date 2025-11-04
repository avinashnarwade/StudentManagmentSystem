from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import Combobox
import os
import openpyxl
from openpyxl import Workbook
import pathlib
from PIL import Image, ImageTk

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file_path = "Book1.xlsx"
file = pathlib.Path(file_path)

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Registration No."
    sheet["B1"] = "Name"
    sheet["C1"] = "Class"
    sheet["D1"] = "Gender"
    sheet["E1"] = "DOB"
    sheet["F1"] = "Date Of Registration"
    sheet["G1"] = "Religion"
    sheet["H1"] = "Skill"
    sheet["I1"] = "Father Name"
    sheet["J1"] = "Mother Name"
    sheet["K1"] = "Father's Occupation"
    sheet["L1"] = "Mother's Occupation"
    file.save(file_path)

########################## Exit window ################################
def exit_app():
    root.destroy()

######################### Show image ####################################
def showimage():
    global filename 
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select image file",
        filetype=(("JPG File", "*.jpg"),
                  ("PNG File", "*.png"),
                  ("All files", "*.*")))
    
    if filename:
        try:
            img = Image.open(filename)
            resized_image = img.resize((190, 190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except Exception as e:
            messagebox.showerror("Error", "The selected file could not be opened as an image.\nPlease select a valid image file.")
            print(e)

##################### Registration no. ####################
def registration_no():
    file = openpyxl.load_workbook(file_path)
    sheet = file.active 
    row = sheet.max_row 
    
    max_row_value = sheet.cell(row=row, column=1).value
    
    try:
        Registration.set(max_row_value + 1)
    except TypeError:
        Registration.set(1)

######################## Clear ######################
def clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")
    Gender.set('')

    registration_no()
    saveButton.config(state='normal')
    
    img1 = PhotoImage(file="images/upload photo.png")
    lbl.config(image=img1)
    lbl.image = img1
    
    img = ""

################Save#############
def save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    G1 = Gender.get()
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()
    
    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("Error", "All fields are required")
        return
    else:
        file = openpyxl.load_workbook(file_path)
        sheet = file.active 
        row = sheet.max_row + 1
        
        sheet.cell(column=1, row=row, value=R1)
        sheet.cell(column=2, row=row, value=N1)
        sheet.cell(column=3, row=row, value=C1)
        sheet.cell(column=4, row=row, value=G1)
        sheet.cell(column=5, row=row, value=D2)
        sheet.cell(column=6, row=row, value=D1)
        sheet.cell(column=7, row=row, value=Re1)
        sheet.cell(column=8, row=row, value=S1)
        sheet.cell(column=9, row=row, value=fathername)
        sheet.cell(column=10, row=row, value=mothername)
        sheet.cell(column=11, row=row, value=F1)
        sheet.cell(column=12, row=row, value=M1)
        
        file.save("Book1.xlsx")
        
        try:
            if filename:
                os.makedirs("Student Image", exist_ok=True)
                img.save("Student Image/" + str(R1) + ".jpg")
        except Exception as e:
            messagebox.showinfo("Info", "Profile picture is not available or could not be saved.")
            print(e)
            
        messagebox.showinfo("Info", "Successfully entered data!")
        clear()
        registration_no()

# Gender selection function
def selection():
    value = radio.get()
    if value == 1:
        Gender.set("Male")
    else:
        Gender.set("Female")

# Top frames
Label(root, text="Email: avinashnarwade872@gmail.com", width=30, height=3, bg="#f0687c", anchor="e").pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=30, height=2, bg="#c36464", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X, pady=10)

# Search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)
imageicon3 = PhotoImage(file="images/search.png")
Srch = Button(root, text="Search", command=None, image=imageicon3, width=123, bg="#68ddfa", font="arial 13 bold")
Srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file="images/Layer 4.png")
Update_button = Button(root, image=imageicon4, bg="#c36464")
Update_button.place(x=110, y=66)

# Registration and date
Label(root, text="Registration No:", font="arial 13", fg="#fff", bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg="#fff", bg=background).place(x=500, y=150)

Registration = StringVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=560, y=150)

# Set the current date in the Date entry
today = date.today()
d1 = today.strftime("%d/%m/%Y")
Date.set(d1)

# Student details frame
obj = LabelFrame(root, text="Student's details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date Of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)


DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=160, y=100)

Gender = StringVar()

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font="arial 10")
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=20, font="arial 10")
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="Roboto 10", width=17, state="readonly")
Class.place(x=630, y=50)
Class.set("Select Class")

Label(obj, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

# Parent's details frame
obj2 = LabelFrame(root, text="Parent's details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="arial 10")
FO_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

M_Name = StringVar()
M_entry = Entry(obj2, textvariable=M_Name, width=20, font="arial 10")
M_entry.place(x=630, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="arial 10")
MO_entry.place(x=630, y=100)

# Image frame
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)


img = PhotoImage(file="images/upload photo.png")
lbl = Label(f, bg="#fff", image=img)
lbl.place(x=0, y=0)

# Buttons
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)
saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=save)
saveButton.place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="lightgrey", command=exit_app).place(x=1000, y=610)

root.mainloop()


# root.mainloop(major("Book1.xlsx"))
